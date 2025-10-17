from pathlib import Path
from typing import List, Dict, Any
from openpyxl import load_workbook
from .base_extractor import BaseExtractor


class XlsxExtractor(BaseExtractor):
    """Extractor converting each worksheet into a 'table' block.
    Cell values are converted to strings.
    """

    def extract_blocks(self, path: Path) -> List[Dict[str, Any]]:
        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(f"File does not exist: {path}")
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
        blocks: List[Dict[str, Any]] = []
        for ws in wb.worksheets:
            rows: List[List[str]] = []
            for row in ws.iter_rows(values_only=True):
                rows.append(["" if v is None else str(v) for v in row])
            # Add table block only if worksheet is not empty
            if any(any(cell for cell in r) for r in rows):
                blocks.append({"type": "table", "table": rows, "sheet": ws.title})
        return blocks
